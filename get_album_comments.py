import html
import os
import sys
from contextlib import suppress

from openpyxl import Workbook

from config import tokens
import utils

MAX_COMMENTS = 100


def get_album_comments(album_link):
    page_id, album_id = album_link[20:].split('_')
    api = utils.VKConnector(token=tokens[0])
    comments_quantity = api.get_comments(page_id, album_id)['count']
    if comments_quantity > 0:
        wb = Workbook()  # wb is for workbook
        ws = wb.active  # ws is for worksheet
        ws['A1'] = 'Время'
        ws['B1'] = 'Ссылка на фото'
        ws['C1'] = 'Весь текст под фото'
        ws['D1'] = 'Название'
        ws['E1'] = 'Артикул'
        ws['F1'] = 'Цена'
        ws['G1'] = 'Имя комментатора'
        ws['H1'] = 'Страница комментатора'
        ws['I1'] = 'Текст комментария'
        offset = 0
        while offset < comments_quantity:
            comments = api.get_comments(page_id, album_id, offset=offset, count=MAX_COMMENTS)['items']
            users_data = api.get_users_json([c['from_id'] for c in comments])
            images_comments = api.get_images_comments(page_id, album_id)
            texts = api.get_photo_text(page_id, album_id)
            for row, comment in enumerate(comments, start=2+offset):
                date = utils.get_readable_date(comment['date'])
                text = html.unescape(comment['text']).replace('+1', 'плюс 1')
                pic_page = f"https://vk.com/photo{page_id}_{comment['pid']}"
                try:
                    user_page = users_data[comment['from_id']]['user_page']
                    name = users_data[comment['from_id']]['name']
                except KeyError:
                    uid = comment['from_id']
                    user_page = f'https://vk.com/id{uid}'
                    try:
                        s = api.get_user_data(uid)[0]
                        name = f"{s['last_name']} {s['first_name']}"
                    except IndexError:
                        name = ''
                    text = texts[comment['pid']].replace('\n', ' ')
                try:
                    image_text = images_comments[comment['pid']]
                except KeyError:
                    image_text = api.get_image_data_by_id(page_id, comment['pid'])['text']
                image_text_split = image_text.split('\n')
                article_flag = 'Артикул: '
                price_flag = 'Цена: '
                product_name = image_text_split[0]
                article = utils.find_by_flag(image_text_split, article_flag)
                with suppress(ValueError):
                    article = int(article)
                price = utils.find_by_flag(image_text_split, price_flag)
                ws[f'A{row}'] = date
                ws[f'B{row}'] = pic_page
                ws[f'B{row}'].hyperlink = pic_page
                ws[f'C{row}'] = image_text.replace('\n', ' ')
                ws[f'D{row}'] = product_name
                ws[f'E{row}'] = article
                ws[f'F{row}'] = price
                ws[f'G{row}'] = name
                ws[f'H{row}'].hyperlink = user_page
                ws[f'I{row}'] = text
            offset += MAX_COMMENTS
        try:
            wb.save(os.path.join(os.path.dirname(sys.argv[0]), 'VK Comments.xlsx'))
        except PermissionError:
            print('Пожалуйста, закройте таблицу с комментариями и запустите скрипт снова.')
            exit(1)
    else:
        print('В этом альбоме ещё нет комментариев. Работа скрипта сейчас будет завершена.')
        exit(1)


if __name__ == '__main__':
    get_album_comments(utils.parse_arguments())
